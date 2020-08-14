import pandas as pd
import sys
import usaddress
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl as xl
import os
from itertools import groupby

# Declaring variables
address_dict = {}
parsed_address_list = []
account_list = []
parsed_address_dict = {}
errored_addresses_dict = {}
column_headers = [
    'AddressNumberPrefix',
    'AddressNumber',
    'AddressNumberSuffix',
    'StreetNamePreModifier',
    'StreetNamePreDirectional',
    'StreetNamePreType',
    'StreetName',
    'StreetNamePostType',
    'StreetNamePostDirectional',
    'StreetNamePostModifier',
    'SecondAddressNumberPrefix',
    'SecondAddressNumber',
    'SecondAddressNumberSuffix',
    'SecondStreetNamePreModifier',
    'SecondStreetNamePreDirectional',
    'SecondStreetNamePreType',
    'SecondStreetName',
    'SecondStreetNamePostType',
    'SecondStreetNamePostDirectional',
    'SubaddressType',
    'SubaddressIdentifier',
    'BuildingName',
    'OccupancyType',
    'OccupancyIdentifier',
    'CornerOf',
    'LandmarkName',
    'PlaceName',
    'StateName',
    'ZipCode',
    'ZipPlus4',
    'CountryName',
    'USPSBoxType',
    'USPSBoxID',
    'USPSBoxGroupType',
    'USPSBoxGroupID',
    'IntersectionSeparator',
    'Recipient',
    'NotAddress',
]

# Prompt user for file location
wb_dir = str(os.path.join(os.environ['USERPROFILE'], "Documents", "scripts", "Parser", "addresses.xlsx"))
# input("Place the addresses spreadsheet in your \"Documents\\Parser\\\" folder.\nMake sure the first sheet contains the account numbers in the first column and the messy addresses in the second column.\n\nPlease enter the name of the spreadsheet: \n\n==>")


# Create pandas DataFrame from spreadsheet
df = pd.read_excel(wb_dir)

# Define error function
def if_tag_error(parsed_string, i):
    errored_addresses_dict[ str(df["account #"][i]) ] = parsed_string

# Read parsed addresses into dictionary
for i in df.index:
    
    # Error exception
    try:
        address_dict[str(df["account #"][i])] = usaddress.tag(str(df["addr"][i]))
        account_list.append(df['account #'][i])
    
    # If there is an error, catch the parsed text and index value, pass to if_tag_error() function
    except usaddress.RepeatedLabelError as e:
        if_tag_error(e.parsed_string, i)

# New DataFrame from parsed addresses.
df2 = pd.DataFrame.from_dict(address_dict, 'index')

# Write DataFrame to new spreadsheet.
writer = ExcelWriter(str(os.path.join(os.environ['USERPROFILE'], "Documents", "scripts", "Parser", "addresses.xlsx")))
df2.to_excel(writer, 'output', index=True)
writer.save()


# Load output workbook for writing
output_wb = load_workbook(filename = str(os.path.join(os.environ['USERPROFILE'], "Documents", "scripts", "Parser", "addresses.xlsx")))


# Create data sheets for parsed and errored addresses.
parsed_addr_sheet = output_wb.create_sheet("parsed")
error_worksheet = output_wb.create_sheet("errors")

# Create parsed and split up address sheet headers
parsed_addr_sheet['A1'] = str('Account #')
parsed_addr_sheet['B1'] = str('AddressNumberPrefix')
parsed_addr_sheet['C1'] = str('AddressNumber')
parsed_addr_sheet['D1'] = str('AddressNumberSuffix')
parsed_addr_sheet['E1'] = str('StreetNamePreModifier')
parsed_addr_sheet['F1'] = str('StreetNamePreDirectional')
parsed_addr_sheet['G1'] = str('StreetNamePreType')
parsed_addr_sheet['H1'] = str('StreetName')
parsed_addr_sheet['I1'] = str('StreetNamePostType')
parsed_addr_sheet['J1'] = str('StreetNamePostDirectional')
parsed_addr_sheet['K1'] = str('StreetNamePostModifier')
parsed_addr_sheet['L1'] = str('SecondAddressNumberPrefix')
parsed_addr_sheet['M1'] = str('SecondAddressNumber')
parsed_addr_sheet['N1'] = str('SecondAddressNumberSuffix')
parsed_addr_sheet['O1'] = str('SecondStreetNamePreModifier')
parsed_addr_sheet['P1'] = str('SecondStreetNamePreDirectional')
parsed_addr_sheet['Q1'] = str('SecondStreetNamePreType')
parsed_addr_sheet['R1'] = str('SecondStreetName')
parsed_addr_sheet['S1'] = str('SecondStreetNamePostType')
parsed_addr_sheet['T1'] = str('SecondStreetNamePostDirectional')
parsed_addr_sheet['U1'] = str('SubaddressType')
parsed_addr_sheet['V1'] = str('SubaddressIdentifier')
parsed_addr_sheet['W1'] = str('BuildingName')
parsed_addr_sheet['X1'] = str('OccupancyType')
parsed_addr_sheet['Y1'] = str('OccupancyIdentifier')
parsed_addr_sheet['Z1'] = str('CornerOf')
parsed_addr_sheet['AA1'] = str('LandmarkName')
parsed_addr_sheet['AB1'] = str('PlaceName')
parsed_addr_sheet['AC1'] = str('StateName')
parsed_addr_sheet['AD1'] = str('ZipCode')
parsed_addr_sheet['AE1'] = str('ZipPlus4')
parsed_addr_sheet['AF1'] = str('CountryName')
parsed_addr_sheet['AG1'] = str('USPSBoxType')
parsed_addr_sheet['AH1'] = str('USPSBoxID')
parsed_addr_sheet['AI1'] = str('USPSBoxGroupType')
parsed_addr_sheet['AJ1'] = str('USPSBoxGroupID')
parsed_addr_sheet['AK1'] = str('IntersectionSeparator')
parsed_addr_sheet['AL1'] = str('Recipient')
parsed_addr_sheet['AM1'] = str('NotAddress')

row_num = 2 # Starting on row 2

# Loop through address dictionary
for key, value in address_dict.items():
   
    # Account Number
    parsed_addr_sheet.cell(row=row_num, column=1).value = key
    
    # Loop through each tuple in address dictionary
    for key2, value2 in address_dict[key][0].items():
        
        # Checking column header
        parsed_col = int(column_headers.index(str(key2)) + 2)
        parsed_addr_sheet.cell(row=row_num, column=parsed_col).value = str(value2)
    row_num += 1       


# Create address error sheet headers
error_worksheet['A1'] = str('Account #')
error_worksheet['B1'] = str('AddressNumberPrefix')
error_worksheet['C1'] = str('AddressNumber')
error_worksheet['D1'] = str('AddressNumberSuffix')
error_worksheet['E1'] = str('StreetNamePreModifier')
error_worksheet['F1'] = str('StreetNamePreDirectional')
error_worksheet['G1'] = str('StreetNamePreType')
error_worksheet['H1'] = str('StreetName')
error_worksheet['I1'] = str('StreetNamePostType')
error_worksheet['J1'] = str('StreetNamePostDirectional')
error_worksheet['K1'] = str('StreetNamePostModifier')
error_worksheet['L1'] = str('SecondAddressNumberPrefix')
error_worksheet['M1'] = str('SecondAddressNumber')
error_worksheet['N1'] = str('SecondAddressNumberSuffix')
error_worksheet['O1'] = str('SecondStreetNamePreModifier')
error_worksheet['P1'] = str('SecondStreetNamePreDirectional')
error_worksheet['Q1'] = str('SecondStreetNamePreType')
error_worksheet['R1'] = str('SecondStreetName')
error_worksheet['S1'] = str('SecondStreetNamePostType')
error_worksheet['T1'] = str('SecondStreetNamePostDirectional')
error_worksheet['U1'] = str('SubaddressType')
error_worksheet['V1'] = str('SubaddressIdentifier')
error_worksheet['W1'] = str('BuildingName')
error_worksheet['X1'] = str('OccupancyType')
error_worksheet['Y1'] = str('OccupancyIdentifier')
error_worksheet['Z1'] = str('CornerOf')
error_worksheet['AA1'] = str('LandmarkName')
error_worksheet['AB1'] = str('PlaceName')
error_worksheet['AC1'] = str('StateName')
error_worksheet['AD1'] = str('ZipCode')
error_worksheet['AE1'] = str('ZipPlus4')
error_worksheet['AF1'] = str('CountryName')
error_worksheet['AG1'] = str('USPSBoxType')
error_worksheet['AH1'] = str('USPSBoxID')
error_worksheet['AI1'] = str('USPSBoxGroupType')
error_worksheet['AJ1'] = str('USPSBoxGroupID')
error_worksheet['AK1'] = str('IntersectionSeparator')
error_worksheet['AL1'] = str('Recipient')
error_worksheet['AM1'] = str('NotAddress')

row_num = 2 # Starting on row 2

# Correct orientation of tuples in error address dictionary
err_address = {}
for key, value in errored_addresses_dict.items():
    err_address[key] = [(y, x) for x, y in value]

# Loop through errored addresses dictionary and place them into the worksheet on the "errors"
for key, value in err_address.items():
    
    # Merge multiples of same label
    [(matchkey,) + tuple(elem for _, elem in group) for matchkey, group in groupby(err_address[key], lambda pair: pair[0])]

    # Account numbers for error addresses
    error_worksheet.cell(row=row_num, column=1).value = key
    
    # Loop through each account # tuple of address pieces
    for i in range(len(err_address[key])):
        
        # Match to column header
        col = int(column_headers.index(str(err_address[key][i][0])) + 2)

        # Set value of cell
        error_worksheet.cell(row=row_num, column=col).value = str(err_address[key][i][1])

    row_num += 1
output_wb.save(str(os.path.join(os.environ['USERPROFILE'], "Documents", "scripts", "Parser", "addresses.xlsx")))
